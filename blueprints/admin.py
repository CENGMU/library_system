import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from openpyxl import load_workbook
from extensions import db
from models import User, Book, Category, BorrowRecord
from utils import admin_required

admin_bp = Blueprint("admin", __name__)

@admin_bp.get("/panel")
@login_required
@admin_required
def panel():
    total_users = User.query.count()
    total_books = Book.query.count()
    borrowing = BorrowRecord.query.filter_by(status="borrowing").count()
    overdue = sum(1 for r in BorrowRecord.query.filter_by(status="borrowing").all() if r.is_overdue)

    top = (
        db.session.query(Book.title, db.func.count(BorrowRecord.id).label("cnt"))
        .join(BorrowRecord, BorrowRecord.book_id == Book.id)
        .group_by(Book.id)
        .order_by(db.desc("cnt"))
        .limit(5)
        .all()
    )

    return render_template(
        "admin_panel.html",
        total_users=total_users,
        total_books=total_books,
        borrowing=borrowing,
        overdue=overdue,
        top=top
    )

@admin_bp.get("/books")
@login_required
@admin_required
def books_admin():
    q = (request.args.get("q") or "").strip()
    cat = (request.args.get("cat") or "").strip()

    query = Book.query
    if q:
        like = f"%{q}%"
        query = query.filter((Book.title.like(like)) | (Book.author.like(like)) | (Book.isbn.like(like)))

    if cat and cat.isdigit():
        query = query.filter(Book.category_id == int(cat))

    page = int(request.args.get("page", 1))
    per_page = 10
    pagination = query.order_by(Book.id.desc()).paginate(page=page, per_page=per_page, error_out=False)

    categories = Category.query.order_by(Category.name.asc()).all()
    return render_template("admin_books.html", pagination=pagination, categories=categories, q=q, cat=cat)

@admin_bp.post("/books/add")
@login_required
@admin_required
def book_add():
    title = (request.form.get("title") or "").strip()
    author = (request.form.get("author") or "").strip()
    isbn = (request.form.get("isbn") or "").strip()
    publisher = (request.form.get("publisher") or "").strip()
    publish_year = (request.form.get("publish_year") or "").strip()
    tags = (request.form.get("tags") or "").strip()
    category_id = request.form.get("category_id") or None
    total = int(request.form.get("total_copies") or 1)

    if not title:
        flash("书名不能为空", "danger")
        return redirect(url_for("admin.books_admin"))

    cat_obj = None
    if category_id and str(category_id).isdigit():
        cat_obj = Category.query.get(int(category_id))

    book = Book(
        title=title,
        author=author,
        isbn=isbn,
        publisher=publisher,
        publish_year=publish_year,
        tags=tags,
        category=cat_obj,
        total_copies=total,
        available_copies=total
    )
    db.session.add(book)
    db.session.commit()

    flash("新增图书成功", "success")
    return redirect(url_for("admin.books_admin"))

@admin_bp.post("/books/delete/<int:book_id>")
@login_required
@admin_required
def book_delete(book_id: int):
    book = Book.query.get_or_404(book_id)
    db.session.delete(book)
    db.session.commit()
    flash("删除成功", "success")
    return redirect(url_for("admin.books_admin"))

@admin_bp.get("/categories")
@login_required
@admin_required
def categories_admin():
    cats = Category.query.order_by(Category.name.asc()).all()
    return render_template("admin_categories.html", cats=cats)

@admin_bp.post("/categories/add")
@login_required
@admin_required
def category_add():
    name = (request.form.get("name") or "").strip()
    if not name:
        flash("分类名不能为空", "danger")
        return redirect(url_for("admin.categories_admin"))
    if Category.query.filter_by(name=name).first():
        flash("分类已存在", "warning")
        return redirect(url_for("admin.categories_admin"))

    db.session.add(Category(name=name))
    db.session.commit()
    flash("新增分类成功", "success")
    return redirect(url_for("admin.categories_admin"))

@admin_bp.post("/categories/delete/<int:cat_id>")
@login_required
@admin_required
def category_delete(cat_id: int):
    cat = Category.query.get_or_404(cat_id)
    db.session.delete(cat)
    db.session.commit()
    flash("删除分类成功", "success")
    return redirect(url_for("admin.categories_admin"))

@admin_bp.get("/users")
@login_required
@admin_required
def users_admin():
    users = User.query.order_by(User.id.desc()).all()
    return render_template("admin_users.html", users=users)

@admin_bp.post("/users/toggle/<int:user_id>")
@login_required
@admin_required
def user_toggle(user_id: int):
    u = User.query.get_or_404(user_id)
    if u.role == "admin":
        u.role = "user"
    else:
        u.role = "admin"
    db.session.commit()
    flash("已切换角色", "success")
    return redirect(url_for("admin.users_admin"))

@admin_bp.get("/excel/template")
@login_required
@admin_required
def excel_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "books"
    ws.append(["title", "author", "isbn", "publisher", "publish_year", "category", "tags", "total_copies"])
    ws.append(["三体", "刘慈欣", "9787536692930", "重庆出版社", "2008", "科幻", "科幻;经典", 5])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="books_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.post("/excel/import")
@login_required
@admin_required
def excel_import():
    f = request.files.get("file")
    if not f:
        flash("请选择Excel文件", "danger")
        return redirect(url_for("admin.books_admin"))

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        ok, fail = 0, 0

        for r in rows:
            title = (r[0] or "").strip() if r[0] else ""
            if not title:
                fail += 1
                continue

            author = (r[1] or "").strip() if r[1] else ""
            isbn = (str(r[2]).strip() if r[2] else "")
            publisher = (r[3] or "").strip() if r[3] else ""
            publish_year = (str(r[4]).strip() if r[4] else "")
            cat_name = (r[5] or "").strip() if r[5] else ""
            tags = (r[6] or "").strip() if r[6] else ""
            total = int(r[7] or 1)

            cat_obj = None
            if cat_name:
                cat_obj = Category.query.filter_by(name=cat_name).first()
                if not cat_obj:
                    cat_obj = Category(name=cat_name)
                    db.session.add(cat_obj)
                    db.session.flush()

            book = Book(
                title=title,
                author=author,
                isbn=isbn,
                publisher=publisher,
                publish_year=publish_year,
                category=cat_obj,
                tags=tags,
                total_copies=total,
                available_copies=total
            )
            db.session.add(book)
            ok += 1

        db.session.commit()
        flash(f"导入完成：成功 {ok} 条，失败 {fail} 条", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"导入失败：{e}", "danger")

    return redirect(url_for("admin.books_admin"))
